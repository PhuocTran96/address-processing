import pandas as pd
import unidecode
import re
import io
from openpyxl.styles import PatternFill

# Đọc dữ liệu từ file Excel
def process_addresses(uploaded_file):
    addresses_df = pd.read_excel(uploaded_file, sheet_name="raw")
    database_df = pd.read_excel(uploaded_file, sheet_name="database")


    # Hàm bỏ dấu tiếng Việt
    def remove_accents(text):
        if pd.isna(text) or not isinstance(text, str):
            return text
        return unidecode.unidecode(str(text))


    def normalize_baria_vungtau(province):
        if not province or not isinstance(province, str):
            return province

        province = province.lower().strip()

        # Các biến thể của Bà Rịa - Vũng Tàu
        brvt_variations = [
            "tỉnh bà rịa - vũng tàu", "tỉnh bà rịa vũng tàu", "tỉnh br - vt",
            "bà rịa", "vũng tàu", "bà rịa vũng tàu", "bà rịa - vũng tàu", "vùng tàu",
            "ba ria", "vung tau", "ba ria vung tau", "ba ria - vung tau"
        ]

        if any(variation in province for variation in brvt_variations) or province in brvt_variations:
            return "bà rịa - vũng tàu"

        return province


    def find_province_first(address):
        """
        Tìm tỉnh/thành phố trong địa chỉ bằng cách so khớp với danh sách tỉnh trong database
        """
        if not isinstance(address, str):
            return None

        # Lấy danh sách tỉnh/thành phố từ database
        provinces = database_df["Tỉnh/Thành phố"].dropna().unique()

        # Chuẩn hóa địa chỉ để tìm kiếm
        normalized_address = address.lower()

        # Đặc biệt xử lý trường hợp "thừa thiên huế"
        if re.search(r'th[ưừ][aà]\s*thi[eê]n\s*hu[êếeé]', normalized_address):
            return "Thừa Thiên - Huế"

        # Tìm tỉnh/thành phố trong địa chỉ
        for province in provinces:
            if province.lower() in normalized_address:
                return province

        # Xử lý các trường hợp đặc biệt của Hồ Chí Minh
        hcmc_variations = [
            "tp hcm", "tp.hcm", "tphcm", "tp. hcm", "hcm", "chm", "tpchm",
            "tp ho chi minh", "tp. ho chi minh", "ho chi minh"
        ]

        normalized_address_no_accent = remove_accents(normalized_address)
        for variation in hcmc_variations:
            if variation in normalized_address_no_accent:
                return "TP Hồ Chí Minh"

        return None


    def preprocess_address(address):
        if not isinstance(address, str):
            return None

        # Handle standalone HCM at the end of address
        address = re.sub(r'\bHCM\b$', 'Hồ Chí Minh', address, flags=re.IGNORECASE)

        # Handle "TP Something HCM" pattern (like "TP Thủ Đức HCM")
        address = re.sub(r'(TP|Tp\.|T\.P\.|Thành phố)\s+([^\s,]+(\s+[^\s,]+)*)\s+HCM\b',
                         r'\1 \2, Hồ Chí Minh', address, flags=re.IGNORECASE)

        # Handle district followed directly by HCM without comma
        address = re.sub(r'(Bình Chánh|Củ Chi|Hóc Môn|Nhà Bè|Cần Giờ|Thủ Đức)\s+(Hồ Chí Minh|HCM|TPHCM|TP HCM)$',
                         r'\1, Hồ Chí Minh', address, flags=re.IGNORECASE)

        # Handle numeric ward patterns in HCMC
        address = re.sub(r'(P\.?\s*(\d+))\s+(Bình Thạnh|Quận \d+|Q\.?\s*\d+)',
                         r'Phường \2, \3', address, flags=re.IGNORECASE)

        # Handle other common patterns
        address = re.sub(r'\bTP\.?\s*HCM\b', 'Hồ Chí Minh', address, flags=re.IGNORECASE)
        address = re.sub(r'\bTPHCM\b', 'Hồ Chí Minh', address, flags=re.IGNORECASE)
        address = re.sub(r'\bTP\.?\s*Hồ\s*Chí\s*Minh\b', 'Hồ Chí Minh', address, flags=re.IGNORECASE)
        address = re.sub(r'\bThành\s*[Pp]hố\s*Hồ\s*Chí\s*Minh\b', 'Hồ Chí Minh', address)

        # Chuẩn hóa dấu phẩy
        address = address.replace(" - ", ", ")
        address = address.replace("-", ", ")

        # Xử lý khoảng trắng thừa
        address = re.sub(r'\s+', ' ', address)
        address = re.sub(r'\s*,\s*', ', ', address)

        # Loại bỏ khoảng trắng ở đầu và cuối
        address = address.strip()

        return address


    # Hàm nhận diện các đơn vị hành chính
    def identify_admin_units(address):
        # Các từ khóa để nhận diện tỉnh/thành phố
        province_keywords = [
            "TP", "Thành phố", "Tỉnh", "Tp.", "T.P", "Tp", "TPHCM", "Tphcm", "HCM",
            "Hà Nội", "Hồ Chí Minh", "Đà Nẵng", "Cần Thơ", "Hải Phòng", "Huế"
        ]

        # Các từ khóa để nhận diện quận/huyện
        district_keywords = [
            "Quận", "Huyện", "Thị xã", "TX.", "TX", "Q.", "Q", "H.", "H"
        ]

        # Các từ khóa để nhận diện phường/xã
        ward_keywords = [
            "Phường", "Xã", "Thị trấn", "TT.", "TT", "P.", "P", "X.", "X",
            "Khu phố", "KP", "Ấp", "Thôn", "Tổ"
        ]

        province = None
        district = None
        ward = None
        detail = None

        # Tìm kiếm trong địa chỉ
        words = address.split()

        # Xử lý trường hợp không có dấu phẩy
        if ", " not in address and len(words) >= 3:
            # Tìm các từ khóa trong địa chỉ
            for i, word in enumerate(words):
                if any(kw in word for kw in province_keywords) and i < len(words) - 1:
                    province = words[i + 1]
                elif any(kw in word for kw in district_keywords) and i < len(words) - 1:
                    district = words[i + 1]
                elif any(kw in word for kw in ward_keywords) and i < len(words) - 1:
                    ward = words[i + 1]

        return province, district, ward, detail


    # Hàm tách địa chỉ thành 3 cấp
    def split_address(address):
        if not isinstance(address, str):
            return None, None, None, None

        # Tiền xử lý địa chỉ
        address = preprocess_address(address)

        # Xử lý đặc biệt cho Bà Rịa - Vũng Tàu
        brvt_pattern = r'(.*?)(?:,\s*)?((?:Thành phố|TP\.?|T\.P\.?|Thị xã|TX\.?|Huyện)\s+([^,]+))(?:,\s*)?(Bà Rịa|Bà Rịa - Vũng Tàu|Vũng Tàu)$'
        brvt_match = re.search(brvt_pattern, address, re.IGNORECASE)

        if brvt_match and normalize_baria_vungtau(brvt_match.group(4)):
            detail_and_ward = brvt_match.group(1).strip() if brvt_match.group(1) else None
            district = brvt_match.group(3).strip()
            province = "Bà Rịa - Vũng Tàu"

            # Tách ward từ detail nếu có
            if detail_and_ward:
                ward_pattern = r'(.*?)(?:,\s*)?([^,]+)$'
                ward_match = re.search(ward_pattern, detail_and_ward)
                if ward_match:
                    detail = ward_match.group(1).strip() if ward_match.group(1) else None
                    ward = ward_match.group(2).strip()
                    return province, district, ward, detail
                else:
                    return province, district, detail_and_ward, None
            else:
                return province, district, None, None

        # Xử lý trường hợp đặc biệt "thành phố Vũng Tàu, tỉnh Bà Rịa - Vũng Tàu"
        vungtau_pattern = r'(.*?)(?:,\s*)?(?:thành phố|tp\.?)\s+vũng\s+tàu(?:,\s*)?(?:tỉnh)?\s+bà\s+rịa(?:\s*-\s*vũng\s+tàu)?'
        vungtau_match = re.search(vungtau_pattern, address.lower())

        if vungtau_match:
            detail_and_ward = vungtau_match.group(1).strip() if vungtau_match.group(1) else None
            district = "Vũng Tàu"
            province = "Bà Rịa - Vũng Tàu"

            # Tách ward từ detail nếu có
            if detail_and_ward and "," in detail_and_ward:
                parts = detail_and_ward.split(",")
                ward = parts[-1].strip()
                detail = ", ".join(parts[:-1]).strip()
                return province, district, ward, detail
            else:
                return province, district, detail_and_ward, None

        # Xử lý các trường hợp đặc biệt khi địa chỉ chứa "Bà Rịa" hoặc "Vũng Tàu" nhưng không theo mẫu trên
        if re.search(r'bà\s*rịa|vũng\s*tàu', address.lower()):
            parts = address.split(", ")

            # Xác định province trước
            province = "Bà Rịa - Vũng Tàu"

            # Tìm district trong các phần còn lại
            district = None
            ward = None
            detail = None

            # Danh sách các district của Bà Rịa - Vũng Tàu
            brvt_districts = ["bà rịa", "vũng tàu", "châu đức", "đất đỏ", "long điền", "côn đảo", "xuyên mộc", "phú mỹ"]

            # Tìm district trong các phần
            for i, part in enumerate(parts):
                part_lower = part.lower()
                if any(district_name in part_lower for district_name in brvt_districts):
                    # Tránh nhầm lẫn "Bà Rịa" và "Vũng Tàu" là district khi chúng là một phần của tên tỉnh
                    if "bà rịa" in part_lower and "vũng tàu" in address.lower():
                        continue
                    if "vũng tàu" in part_lower and "bà rịa" in address.lower():
                        continue

                    district = part

                    # Ward có thể là phần trước district
                    if i > 0:
                        ward = parts[i - 1]

                    # Detail là các phần còn lại
                    if i > 1:
                        detail = ", ".join(parts[:i - 1])

                    break

            if district:
                return province, district, ward, detail

        # Tìm tỉnh/thành phố trước
        province = find_province_first(address)

        # Nếu tìm thấy tỉnh/thành phố, tiếp tục tách các thành phần khác
        if province:
            # Thử tách theo dấu phẩy
            parts = address.split(", ")

            # Xử lý các trường hợp có nhiều hơn 3 phần
            if len(parts) > 3:
                ward = parts[-3]
                district = parts[-2]
                detail = ", ".join(parts[:-3]).rstrip()  # Join all remaining parts as detail
                return province, district, ward, detail

            # Xử lý các trường hợp có đủ 3 phần
            elif len(parts) == 3:
                ward = parts[0]
                district = parts[1]
                return province, district, ward, None  # No detail

            # Xử lý các trường hợp chỉ có 2 phần
            elif len(parts) == 2:
                district = parts[0]
                return province, district, None, None  # No ward, no detail

            # Trường hợp không có dấu phẩy hoặc chỉ có 1 phần
            else:
                # Thử nhận diện các đơn vị hành chính
                _, district, ward, detail = identify_admin_units(address)
                return province, district, ward, detail

        # Thử tách theo dấu phẩy
        parts = address.split(", ")

        # Xử lý các trường hợp có nhiều hơn 3 phần
        if len(parts) > 3:
            ward = parts[-3]
            district = parts[-2]
            province = parts[-1]
            detail = ", ".join(parts[:-3]).rstrip()  # Join all remaining parts as detail
            return province, district, ward, detail

        # Xử lý các trường hợp có đủ 3 phần
        elif len(parts) == 3:
            ward = parts[0]
            district = parts[1]
            province = parts[2]
            return province, district, ward, None  # No detail

        # Xử lý các trường hợp chỉ có 2 phần
        elif len(parts) == 2:
            district = parts[0]
            province = parts[1]
            return province, district, None, None  # No ward, no detail

        # Trường hợp không có dấu phẩy hoặc chỉ có 1 phần
        else:
            # Thử nhận diện các đơn vị hành chính
            province, district, ward, detail = identify_admin_units(address)

            # Nếu không nhận diện được, xử lý theo không gian
            if not any([province, district, ward]):
                words = address.split()
                if len(words) >= 3:
                    # Giả định 3 từ cuối lần lượt là phường/xã, quận/huyện, tỉnh/thành phố
                    ward = ' '.join(words[:-2])
                    district = words[-2]
                    province = words[-1]
                    return province, district, ward, None
                elif len(words) == 2:
                    district = words[0]
                    province = words[1]
                    return province, district, None, None

            return address, None, None, None  # Trả về toàn bộ địa chỉ nếu không thể phân tích


    def normalize_province(province):
        if not province or not isinstance(province, str):
            return province

        # Kiểm tra Bà Rịa - Vũng Tàu
        brvt_normalized = normalize_baria_vungtau(province)
        if brvt_normalized != province.lower().strip():
            return brvt_normalized

        province = province.lower().strip()

        # Chuẩn hóa Thừa Thiên Huế
        if re.search(r'th[ưừ][aà]\s*thi[eê]n\s*hu[êếeé]', province):
            return "thừa thiên - huế"

        # Các chuẩn hóa khác...
        if province.startswith("tỉnh "):
            return province[5:].strip()
        if province.startswith("tp ") or province.startswith("tp. "):
            return province[3:].strip() if province.startswith("tp ") else province[4:].strip()
        if province.startswith("thành phố "):
            return province[10:].strip()

        return province


    # Chuẩn hóa tên quận/huyện
    def normalize_district(district):
        if district and isinstance(district, str):
            # Loại bỏ tiền tố nếu có
            district = district.lower().strip()
            prefixes = ["quận 0", "quận ", "huyện ", "thị xã ", "tx. ", "tx ", "tp. ", "tp ", "thành phố "]
            for prefix in prefixes:
                if district.startswith(prefix):
                    return district[len(prefix):].strip()
            return district
        return district


    # Chuẩn hóa tên phường/xã
    def normalize_ward(ward):
        if ward and isinstance(ward, str):
            # Loại bỏ tiền tố nếu có
            ward = ward.lower().strip()
            prefixes = ["p.", "p. ", "phường ", "xã ", "thị trấn ", "tt. ", "tt ", "khu phố ", "kp ", "ấp ", "thôn ",
                        "tổ ", "p", "p "]
            for prefix in prefixes:
                if ward.startswith(prefix):
                    return ward[len(prefix):].strip()
            return ward
        return ward

    # Tạo DataFrame mới với các cột đã tách
    result_df = pd.DataFrame(addresses_df["Address"].tolist(), columns=["Address"])
    result_df[["Province/City", "District", "Ward", "Detail"]] = result_df["Address"].apply(
        lambda x: pd.Series(split_address(x))
    )


    # Áp dụng hàm chuẩn hóa
    result_df["Normalized Province"] = result_df["Province/City"].apply(normalize_province)
    result_df["Normalized District"] = result_df["District"].apply(normalize_district)
    result_df["Normalized Ward"] = result_df["Ward"].apply(normalize_ward)

    # Tạo phiên bản không dấu
    result_df["Province No Accent"] = result_df["Normalized Province"].apply(remove_accents)
    result_df["District No Accent"] = result_df["Normalized District"].apply(remove_accents)
    result_df["Ward No Accent"] = result_df["Normalized Ward"].apply(remove_accents)

    # Tạo các cột mã số
    result_df["Province Code"] = None
    result_df["District Code"] = None
    result_df["Ward Code"] = None

    # Chuẩn bị database để tìm kiếm
    # Tạo các cột chuẩn hóa và không dấu trong database
    database_df["Normalized Province"] = database_df["Tỉnh/Thành phố"].apply(normalize_province)
    database_df["Normalized District"] = database_df["Quận/Huyện"].apply(normalize_district)
    database_df["Normalized Ward"] = database_df["Phường/Xã"].apply(normalize_ward)
    database_df["Province No Accent"] = database_df["Normalized Province"].apply(remove_accents)
    database_df["District No Accent"] = database_df["Normalized District"].apply(remove_accents)
    database_df["Ward No Accent"] = database_df["Normalized Ward"].apply(remove_accents)

    # Tạo từ điển ánh xạ tên tỉnh -> mã tỉnh
    province_code_map = {}
    for _, row in database_df.dropna(subset=["Tỉnh/Thành phố", "Mã Tỉnh/Thành phố"]).iterrows():
        province_name = row["Normalized Province"]
        province_no_accent = row["Province No Accent"]
        province_code = row["Mã Tỉnh/Thành phố"]
        if pd.notna(province_name) and pd.notna(province_code):
            if province_name not in province_code_map:
                province_code_map[province_name] = province_code
            if province_no_accent not in province_code_map:
                province_code_map[province_no_accent] = province_code

    # Tìm mã tỉnh/thành phố, quận/huyện, phường/xã
    for index, row in result_df.iterrows():
        # Tìm mã tỉnh/thành phố
        province_name = row["Normalized Province"]
        province_no_accent = row["Province No Accent"]
        if province_name in province_code_map:
            province_code = province_code_map[province_name]
        elif province_no_accent in province_code_map:
            province_code = province_code_map[province_no_accent]
        else:
            continue

        result_df.at[index, "Province Code"] = province_code

        # Tìm mã quận/huyện dựa trên mã tỉnh
        if pd.notna(province_code):
            province_code_str = str(int(province_code))  # Chuyển về dạng số nguyên rồi thành chuỗi để loại bỏ .0
            district_name = row["Normalized District"]
            district_no_accent = row["District No Accent"]
            if pd.isna(district_name):
                continue

            # Tìm tất cả quận/huyện có mã bắt đầu bằng mã tỉnh
            district_pattern = f"^{province_code_str}"
            matching_districts = database_df[database_df["Mã Quận/Huyện"].astype(str).str.match(district_pattern)]
            district_code = None

            # Tìm quận/huyện phù hợp
            for _, district_row in matching_districts.iterrows():
                db_district_name = district_row["Normalized District"]
                db_district_no_accent = district_row["District No Accent"]
                if (pd.notna(db_district_name) and db_district_name == district_name) or \
                        (pd.notna(db_district_no_accent) and db_district_no_accent == district_no_accent):
                    district_code = district_row["Mã Quận/Huyện"]
                    result_df.at[index, "District Code"] = district_code
                    break

            # Tìm mã phường/xã dựa trên mã quận/huyện
            if pd.notna(district_code) and pd.notna(row["Normalized Ward"]):
                district_code_str = str(int(district_code))
                ward_name = row["Normalized Ward"]
                ward_no_accent = row["Ward No Accent"]

                # Tìm tất cả phường/xã có mã bắt đầu bằng mã quận/huyện
                ward_pattern = f"^{district_code_str}"
                matching_wards = database_df[database_df["Mã Phường/Xã"].astype(str).str.match(ward_pattern)]

                # Tìm phường/xã phù hợp
                for _, ward_row in matching_wards.iterrows():
                    db_ward_name = ward_row["Normalized Ward"]
                    db_ward_no_accent = ward_row["Ward No Accent"]
                    if (pd.notna(db_ward_name) and db_ward_name == ward_name) or \
                            (pd.notna(db_ward_no_accent) and db_ward_no_accent == ward_no_accent):
                        ward_code = ward_row["Mã Phường/Xã"]
                        result_df.at[index, "Ward Code"] = ward_code
                        break

    # Thêm sau phần tìm mã Ward Code

    # Tạo từ điển ánh xạ mã -> tên chuẩn
    province_name_map = {}
    district_name_map = {}
    ward_name_map = {}

    # Tạo mapping từ mã code đến tên chuẩn
    for _, row in database_df.dropna(subset=["Mã Tỉnh/Thành phố", "Tỉnh/Thành phố"]).iterrows():
        province_code = row["Mã Tỉnh/Thành phố"]
        province_name = row["Tỉnh/Thành phố"]
        if pd.notna(province_code) and pd.notna(province_name):
            province_name_map[str(int(province_code))] = province_name

    for _, row in database_df.dropna(subset=["Mã Quận/Huyện", "Quận/Huyện"]).iterrows():
        district_code = row["Mã Quận/Huyện"]
        district_name = row["Quận/Huyện"]
        if pd.notna(district_code) and pd.notna(district_name):
            district_name_map[str(int(district_code))] = district_name

    for _, row in database_df.dropna(subset=["Mã Phường/Xã", "Phường/Xã"]).iterrows():
        ward_code = row["Mã Phường/Xã"]
        ward_name = row["Phường/Xã"]
        if pd.notna(ward_code) and pd.notna(ward_name):
            ward_name_map[str(int(ward_code))] = ward_name

    # Cập nhật tên chuẩn dựa trên mã code
    for index, row in result_df.iterrows():
        # Cập nhật tên tỉnh/thành phố
        if pd.notna(row["Province Code"]):
            province_code_str = str(int(row["Province Code"]))
            if province_code_str in province_name_map:
                result_df.at[index, "Province/City"] = province_name_map[province_code_str]

        # Cập nhật tên quận/huyện
        if pd.notna(row["District Code"]):
            district_code_str = str(int(row["District Code"]))
            if district_code_str in district_name_map:
                result_df.at[index, "District"] = district_name_map[district_code_str]

        # Cập nhật tên phường/xã
        if pd.notna(row["Ward Code"]):
            ward_code_str = str(int(row["Ward Code"]))
            if ward_code_str in ward_name_map:
                result_df.at[index, "Ward"] = ward_name_map[ward_code_str]

    # Lưu kết quả vào file Excel mới
    result_df['Check'] = (result_df[['Province Code', 'District Code', 'Ward Code']].isnull().any(axis=1)
                                     .map({True: "Cần kiểm tra", False: ""}))
    result_df = result_df[["Address", "Detail", "Ward Code", "Ward", "District Code", "District", "Province Code",
                           "Province/City", "Check"]]

    return result_df


def generate_excel(result_df):
    # Tạo BytesIO object để lưu file Excel trong bộ nhớ
    output = io.BytesIO()

    # Sử dụng ExcelWriter với engine là openpyxl
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name="Processed Data")

        # Lấy workbook và worksheet
        workbook = writer.book
        worksheet = writer.sheets["Processed Data"]

        # Áp dụng định dạng
        yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")

        # Tô màu các ô trống trong cột mã
        for row_idx, row in enumerate(result_df.values, start=2):  # Start from 2 to skip header
            for col_idx, value in [(2, row[2]), (4, row[4]), (6, row[6])]:  # Ward Code, District Code, Province Code
                if pd.isna(value):
                    cell = worksheet.cell(row=row_idx, column=col_idx + 1)  # +1 because openpyxl is 1-indexed
                    cell.fill = yellow_fill

    # Đặt con trỏ về đầu file
    output.seek(0)
    return output


